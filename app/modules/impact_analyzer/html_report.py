from __future__ import annotations

from collections import Counter
from pathlib import Path

from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.config import get_settings
from app.modules.impact_analyzer.schemas import AnalysisResult

_templates = Jinja2Templates(directory=[Path(__file__).parent / "templates", get_settings().root / "app" / "web" / "templates"])


def _impact_areas(result: AnalysisResult) -> list[dict]:
    return [
        {"feature": item.feature, "related_modules": item.related_modules, "impact_area": item.impact_area}
        for item in result.change.change_items
        if item.feature or item.impact_area
    ]


def create_html_report(result: AnalysisResult) -> Path:
    output = get_settings().path("storage.report_dir") / f"regression-{result.analysis_id}.html"
    context = {
        "result": result,
        "counts": Counter(item.impact.value for item in result.decisions),
        "recommended": [item for item in result.decisions if item.recommended],
        "impact_areas": _impact_areas(result),
    }
    html_text = _templates.get_template("report.html").render(context)
    output.write_text(html_text, encoding="utf-8")
    return output


def create_xlsx_export(result: AnalysisResult) -> Path:
    output = get_settings().path("storage.export_dir") / f"regression-{result.analysis_id}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Regression Impact"
    headers = ["TC ID", "Impact", "Confidence", "Evidence Level", "Revision Mark", "Reason", "Specification Reference", "Related Spec Chunk", "Verification Point", "Recommended", "Manual Review"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B87")
    for item in result.decisions:
        sheet.append([item.tc_id, item.impact.value, item.confidence, item.evidence_level.value, item.revision_mark.value, item.reason, item.specification_reference, " | ".join(item.relevant_specifications), " | ".join(item.verification_points), item.recommended, item.manual_review_required])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (18, 12, 12, 16, 14, 48, 30, 35, 35, 14, 14)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(output)
    return output
