from __future__ import annotations

import csv
import html
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.config import get_settings
from app.core.schemas import AnalysisResult


def create_html_report(result: AnalysisResult) -> Path:
    output = get_settings().path("storage.report_dir") / f"regression-{result.analysis_id}.html"
    counts = Counter(item.impact.value for item in result.decisions)
    rows = "".join(f"<tr><td>{html.escape(item.tc_id)}</td><td><span class='impact {item.impact.value}'>{item.impact.value}</span></td><td>{item.confidence:.2f}</td><td>{html.escape(item.reason)}</td><td>{'<br>'.join(map(html.escape, item.relevant_specifications)) or '-'}</td><td>{'<br>'.join(map(html.escape, item.verification_points)) or '-'}</td></tr>" for item in result.decisions if item.recommended)
    manual = "".join(f"<li>{html.escape(item.tc_id)} — {html.escape(item.reason)}</li>" for item in result.decisions if item.manual_review_required) or "<li>없음</li>"
    features = "".join(f"<li>{html.escape(value)}</li>" for value in result.change.changed_features) or "<li>자동 추출 결과 없음</li>"
    document = f"""<!doctype html><html lang='ko'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>AI Regression Impact Analysis Report</title><style>body{{font-family:Arial,'Noto Sans KR',sans-serif;margin:0;background:#f4f7fb;color:#172033}}main{{max-width:1200px;margin:auto;padding:28px}}header{{background:#173b67;color:white;padding:24px;border-radius:14px}}.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin:20px 0}}.card,section{{background:white;border-radius:12px;padding:18px;box-shadow:0 3px 14px #17203312}}table{{width:100%;border-collapse:collapse}}th,td{{padding:10px;border-bottom:1px solid #dde4ee;text-align:left;vertical-align:top}}.table{{overflow:auto}}.impact{{font-weight:bold}}.HIGH{{color:#b42318}}.MEDIUM{{color:#b54708}}.LOW{{color:#175cd3}}.NONE{{color:#667085}}small{{opacity:.8}}@media(max-width:650px){{main{{padding:12px}}th,td{{font-size:13px}}}}</style></head><body><main><header><h1>AI Regression Impact Analysis Report</h1><small>{result.created_at.isoformat()}</small></header><section><h2>1. 분석 정보</h2><p>변경사항: {html.escape(result.change_file)}<br>사양서: {html.escape(result.specification_file)}<br>TC: {html.escape(result.testcase_file)}</p></section><section><h2>2. Change Summary</h2><p>{html.escape(result.change.purpose)}</p><ul>{features}</ul></section><div class='cards'><div class='card'><b>전체 TC</b><h2>{result.total_tc}</h2></div><div class='card'><b>Candidate</b><h2>{result.candidate_tc}</h2></div><div class='card'><b>Recommended</b><h2>{result.recommended_count}</h2></div>{''.join(f'<div class="card"><b>{key}</b><h2>{counts[key]}</h2></div>' for key in ('HIGH','MEDIUM','LOW','NONE'))}</div><section><h2>4. Recommended Regression TC</h2><div class='table'><table><thead><tr><th>TC ID</th><th>Impact</th><th>Confidence</th><th>Reason</th><th>Related Specification</th><th>Verification Point</th></tr></thead><tbody>{rows}</tbody></table></div></section><section><h2>5. Manual Review Required</h2><ul>{manual}</ul></section><section><h2>6. AI 판단 주의사항</h2><p>AI 결과는 후보 추천이며 최종 QA 판단을 대체하지 않습니다. 낮은 신뢰도 또는 참조 근거가 없는 결과는 수동 검토가 필요합니다.</p></section></main></body></html>"""
    output.write_text(document, encoding="utf-8")
    return output


def create_csv_export(result: AnalysisResult) -> Path:
    output = get_settings().path("storage.export_dir") / f"regression-{result.analysis_id}.csv"
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["TC ID", "Impact", "Confidence", "Reason", "Related Specification", "Verification Point", "Recommended", "Manual Review"])
        for item in result.decisions:
            writer.writerow([item.tc_id, item.impact.value, item.confidence, item.reason, " | ".join(item.relevant_specifications), " | ".join(item.verification_points), item.recommended, item.manual_review_required])
    return output


def create_xlsx_export(result: AnalysisResult) -> Path:
    output = get_settings().path("storage.export_dir") / f"regression-{result.analysis_id}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Regression Impact"
    headers = ["TC ID", "Impact", "Confidence", "Reason", "Related Specification", "Verification Point", "Recommended", "Manual Review"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B87")
    for item in result.decisions:
        sheet.append([item.tc_id, item.impact.value, item.confidence, item.reason, " | ".join(item.relevant_specifications), " | ".join(item.verification_points), item.recommended, item.manual_review_required])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (18, 12, 12, 48, 35, 35, 14, 14)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(output)
    return output
