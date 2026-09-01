from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from app.modules.impact_analyzer.schemas import TestCase

ALIASES = {
    "tc_id": ["tcid", "testcaseid", "testid", "oldid", "id", "tcno", "테스트케이스id", "tc번호"],
    "category": ["category", "stccategory", "분류", "구분"],
    "feature": ["feature", "function", "title", "testitem", "기능", "기능명", "requirement"],
    "precondition": ["precondition", "pre-condition", "pre_condition", "사전조건", "전제조건"],
    "step": ["step", "steps", "teststep", "stepdescription", "절차", "시험절차"],
    "expected_result": ["expectedresult", "expectedtestresult", "expectedreuslt", "expected", "기대결과", "예상결과"],
    "result": ["result", "testresult", "finalresult", "최종result", "결과"],
    "remark": ["remark", "remarks", "comment", "comments", "description", "비고", "note"],
}


def _normalize(value: object) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").lower())


def detect_columns(headers: list[object], mapping: dict[str, str] | None = None) -> dict[str, int]:
    normalized = {_normalize(value): index for index, value in enumerate(headers)}
    result: dict[str, int] = {}
    for field, aliases in ALIASES.items():
        configured = (mapping or {}).get(field)
        candidates = ([configured] if configured else []) + aliases
        for candidate in candidates:
            key = _normalize(candidate)
            if key in normalized:
                result[field] = normalized[key]
                break
    if "tc_id" not in result:
        raise ValueError("TC ID 컬럼을 찾을 수 없습니다. 컬럼 매핑을 확인하세요.")
    return result


def suggest_columns(headers: list[object]) -> dict[int, str]:
    """`detect_columns`와 같은 별칭 규칙으로 컬럼 인덱스 -> 필드명 추정치를 반환한다.
    `tc_id`를 못 찾아도 예외를 던지지 않는다 — 수동 매핑 UI에서 드롭다운 기본값을
    미리 채우는 용도라 부분 추정만으로도 유용하다."""
    normalized = {_normalize(value): index for index, value in enumerate(headers)}
    result: dict[int, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            key = _normalize(alias)
            if key in normalized:
                result[normalized[key]] = field
                break
    return result


def preview_workbook(path: Path, max_rows: int = 15) -> dict[str, list[list[str]]]:
    """시트명 -> 상위 max_rows행의 셀 값(문자열) 목록. 수동 컬럼/시트 매핑 UI의 미리보기용."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    preview: dict[str, list[list[str]]] = {}
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, max_rows), values_only=True):
            rows.append(["" if value is None else str(value) for value in row])
        preview[sheet.title] = rows
    return preview


def _find_header(sheet, mapping: dict[str, str] | None) -> tuple[int | None, dict[str, int] | None]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80), values_only=True), start=1):
        try:
            candidate = detect_columns(list(row), mapping)
        except ValueError:
            continue
        if len(candidate) >= 2:
            return row_number, candidate
    return None, None


def parse_testcases(
    path: Path, mapping: dict[str, str] | None = None,
    sheet_name: str | None = None, header_row: int | None = None,
) -> list[TestCase]:
    """`sheet_name`/`header_row`를 둘 다 지정하면 자동 탐지를 건너뛰고 그 위치를 헤더로
    강제 사용한다(수동 매핑 UI 경유) — 이 경우 컬럼을 찾지 못하면 다른 시트로 넘어가지
    않고 바로 실패해 사용자가 원인을 즉시 알 수 있게 한다."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        cases: list[TestCase] = []
        seen: set[str] = set()
        detected_sheet = False
        sheets = [workbook[sheet_name]] if sheet_name and header_row else workbook.worksheets
        for sheet in sheets:
            if sheet_name and header_row:
                header_values = next(iter(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)), None)
                resolved_row = header_row if header_values is not None else None
                columns = detect_columns(list(header_values), mapping) if header_values is not None else None
            else:
                resolved_row, columns = _find_header(sheet, mapping)
            if resolved_row is None or columns is None:
                continue
            detected_sheet = True
            for row in sheet.iter_rows(min_row=resolved_row + 1, values_only=True):
                values = {
                    field: str(row[index] or "").strip()
                    for field, index in columns.items()
                    if index < len(row)
                }
                tc_id = values.get("tc_id", "")
                if not tc_id or tc_id in seen:
                    continue
                seen.add(tc_id)
                cases.append(TestCase(**values))
        if not detected_sheet:
            raise ValueError("TC ID 컬럼을 찾을 수 없습니다. 컬럼 매핑을 확인하세요.")
        return cases
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Excel을 읽을 수 없습니다: {path.name}") from exc
