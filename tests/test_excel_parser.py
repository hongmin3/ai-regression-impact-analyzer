import pytest
from openpyxl import Workbook

from app.parsers.excel_parser import detect_columns, parse_testcases, preview_workbook, suggest_columns


def test_detect_localized_columns():
    columns = detect_columns(["TC 번호", "기능명", "사전 조건", "시험 절차", "기대 결과"])
    assert columns["tc_id"] == 0
    assert columns["expected_result"] == 4


def test_parse_and_deduplicate(tmp_path):
    path = tmp_path / "tc.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC ID", "Feature", "Step", "Expected Result"])
    ws.append(["TC-1", "설정", "저장", "유지"])
    ws.append(["TC-1", "설정", "반복", "유지"])
    wb.save(path)
    cases = parse_testcases(path)
    assert len(cases) == 1
    assert cases[0].tc_id == "TC-1"


def test_unrecognized_headers_fail_auto_detection(tmp_path):
    path = tmp_path / "unrecognized.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["관리번호", "세부기능", "확인절차", "판정기준"])
    ws.append(["CASE-1", "설정", "저장", "유지"])
    wb.save(path)

    with pytest.raises(ValueError, match="TC ID 컬럼을 찾을 수 없습니다"):
        parse_testcases(path)


def test_manual_sheet_and_header_row_bypasses_auto_detection(tmp_path):
    path = tmp_path / "manual.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "본문"
    ws.append(["문서 안내", "", "", ""])  # 1행: 자동 탐지가 못 알아보는 설명 줄
    ws.append(["관리번호", "세부기능", "확인절차", "판정기준"])  # 2행: 실제 헤더(별칭에 없음)
    ws.append(["CASE-1", "설정", "저장", "유지"])
    wb.save(path)

    cases = parse_testcases(
        path,
        mapping={"tc_id": "관리번호", "feature": "세부기능", "step": "확인절차", "expected_result": "판정기준"},
        sheet_name="본문", header_row=2,
    )

    assert len(cases) == 1
    assert cases[0].tc_id == "CASE-1"
    assert cases[0].step == "저장"


def test_manual_mapping_still_raises_when_tc_id_unresolved(tmp_path):
    path = tmp_path / "bad_mapping.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["관리번호", "세부기능"])
    ws.append(["CASE-1", "설정"])
    wb.save(path)

    with pytest.raises(ValueError, match="TC ID 컬럼을 찾을 수 없습니다"):
        parse_testcases(path, mapping={"feature": "세부기능"}, sheet_name="Sheet", header_row=1)


def test_suggest_columns_never_raises_and_returns_partial_guesses():
    guesses = suggest_columns(["관리번호", "Feature", "확인절차"])
    assert guesses == {1: "feature"}  # "관리번호"/"확인절차"는 별칭에 없어 추정되지 않음


def test_preview_workbook_returns_string_rows_per_sheet(tmp_path):
    path = tmp_path / "preview.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, None])
    wb.save(path)

    preview = preview_workbook(path)

    assert preview["Sheet1"][0] == ["A", "B"]
    assert preview["Sheet1"][1] == ["1", ""]
