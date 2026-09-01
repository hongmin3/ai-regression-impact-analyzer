import io
import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.modules.knowledge.router as knowledge_router
from app.core.storage import Storage
from app.main import app


class _FakeSettings:
    """`get_settings().path("storage.testcase_dir")`만 tmp_path로 돌려 실제 data/testcases를
    건드리지 않게 하는 최소 stub. 이 라우트들은 이 키만 사용한다."""

    def __init__(self, testcase_dir: Path) -> None:
        self._testcase_dir = testcase_dir

    def path(self, dotted: str) -> Path:
        assert dotted == "storage.testcase_dir"
        return self._testcase_dir


def _xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _use_fake_testcase_dir(monkeypatch, tmp_path) -> Path:
    testcase_dir = tmp_path / "testcases"
    testcase_dir.mkdir()
    monkeypatch.setattr(knowledge_router, "get_settings", lambda: _FakeSettings(testcase_dir))
    return testcase_dir


def test_register_testcase_with_recognized_headers_registers_directly(monkeypatch, tmp_path):
    storage = Storage(tmp_path / "app.db")
    monkeypatch.setattr(knowledge_router, "storage", storage)
    _use_fake_testcase_dir(monkeypatch, tmp_path)
    content = _xlsx_bytes([["TC ID", "Feature", "Step", "Expected Result"], ["TC-1", "설정", "저장", "유지"]])

    response = TestClient(app).post(
        "/knowledge/testcase",
        files={"file": ("tc.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"product": "VXvue", "version": "1.0"},
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/knowledge"
    docs = storage.active_documents("testcase", "VXvue")
    assert len(docs) == 1
    assert docs[0]["name"] == "tc.xlsx"


def test_register_testcase_with_unrecognized_headers_redirects_to_mapping(monkeypatch, tmp_path):
    storage = Storage(tmp_path / "app.db")
    monkeypatch.setattr(knowledge_router, "storage", storage)
    _use_fake_testcase_dir(monkeypatch, tmp_path)
    content = _xlsx_bytes([["관리번호", "세부기능", "확인절차"], ["CASE-1", "설정", "저장"]])

    response = TestClient(app).post(
        "/knowledge/testcase",
        files={"file": ("odd_headers.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"product": "VXvue", "version": "1.0"},
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"].startswith("/knowledge/testcase/map?")
    assert "odd_headers.xlsx" in response.headers["location"]
    # 자동 등록에 실패했으니 documents 테이블에는 아직 아무것도 없어야 한다.
    assert storage.active_documents("testcase", "VXvue") == []


def test_mapping_page_shows_preview_and_suggested_fields(monkeypatch, tmp_path):
    storage = Storage(tmp_path / "app.db")
    monkeypatch.setattr(knowledge_router, "storage", storage)
    testcase_dir = _use_fake_testcase_dir(monkeypatch, tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["관리번호", "Feature", "확인절차"])
    ws.append(["CASE-1", "설정", "저장"])
    wb.save(testcase_dir / "sample.xlsx")

    response = TestClient(app).get(
        "/knowledge/testcase/map",
        params={"filename": "sample.xlsx", "product": "VXvue", "version": "1.0", "original_name": "sample.xlsx", "header_row": 1},
    )

    assert response.status_code == 200
    assert "관리번호" in response.text
    assert "확인절차" in response.text
    assert "기능명" in response.text  # feature label


def test_mapping_page_rejects_path_traversal_filename(monkeypatch, tmp_path):
    _use_fake_testcase_dir(monkeypatch, tmp_path)
    response = TestClient(app).get(
        "/knowledge/testcase/map",
        params={"filename": "../secrets.txt", "product": "VXvue"},
    )
    assert response.status_code == 400


def test_mapping_page_404s_when_upload_missing(monkeypatch, tmp_path):
    _use_fake_testcase_dir(monkeypatch, tmp_path)
    response = TestClient(app).get(
        "/knowledge/testcase/map",
        params={"filename": "does-not-exist.xlsx", "product": "VXvue"},
    )
    assert response.status_code == 404


def test_submit_mapping_registers_document_with_stored_mapping(monkeypatch, tmp_path):
    storage = Storage(tmp_path / "app.db")
    monkeypatch.setattr(knowledge_router, "storage", storage)
    testcase_dir = _use_fake_testcase_dir(monkeypatch, tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "본문"
    ws.append(["관리번호", "세부기능", "확인절차"])
    ws.append(["CASE-1", "설정", "저장"])
    wb.save(testcase_dir / "sample.xlsx")

    response = TestClient(app).post(
        "/knowledge/testcase/map",
        data={
            "filename": "sample.xlsx", "product": "VXvue", "version": "1.0", "original_name": "sample.xlsx",
            "sheet": "본문", "header_row": "1",
            "tc_id": "관리번호", "feature": "세부기능", "step": "확인절차",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/knowledge"
    docs = storage.active_documents("testcase", "VXvue")
    assert len(docs) == 1
    metadata = json.loads(docs[0]["metadata_json"])
    assert metadata["column_mapping"]["tc_id"] == "관리번호"
    assert metadata["sheet_name"] == "본문"
    assert metadata["header_row"] == 1


def test_submit_mapping_without_valid_tc_id_column_redirects_back_with_error(monkeypatch, tmp_path):
    storage = Storage(tmp_path / "app.db")
    monkeypatch.setattr(knowledge_router, "storage", storage)
    testcase_dir = _use_fake_testcase_dir(monkeypatch, tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["관리번호", "세부기능"])
    ws.append(["CASE-1", "설정"])
    wb.save(testcase_dir / "sample.xlsx")

    response = TestClient(app).post(
        "/knowledge/testcase/map",
        data={
            "filename": "sample.xlsx", "product": "VXvue", "version": "1.0", "original_name": "sample.xlsx",
            "sheet": "Sheet", "header_row": "1", "feature": "세부기능",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"].startswith("/knowledge/testcase/map?")
    assert storage.active_documents("testcase", "VXvue") == []
