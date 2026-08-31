from openpyxl import Workbook

from app.parsers.excel_parser import parse_testcases


def test_parser_finds_vxvue_style_headers_after_cover_sheet(tmp_path):
    path = tmp_path / "vxvue.xlsx"
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover.append(["VXvue Test Case"])
    cases = workbook.create_sheet("Basic Function")
    cases.append(["STC Category", "Old ID", "Function", "Precondition", "Step Description", "Expected Result", "Test Result", "Comments"])
    cases.append(["Basic", "TC_BASIC_001", "Registration", "로그인", "환자를 등록한다.", "환자가 표시된다.", "Pass", "기존 이력 유지"])
    workbook.save(path)

    parsed = parse_testcases(path)

    assert len(parsed) == 1
    assert parsed[0].tc_id == "TC_BASIC_001"
    assert parsed[0].step == "환자를 등록한다."
    assert parsed[0].expected_result == "환자가 표시된다."
    assert parsed[0].result == "Pass"
    assert parsed[0].remark == "기존 이력 유지"


def test_parser_deduplicates_tc_ids_across_sheets(tmp_path):
    path = tmp_path / "duplicate.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["TC_ID", "Step", "Expected Test Result"])
    first.append(["TC-1", "첫 절차", "첫 결과"])
    second = workbook.create_sheet("Second")
    second.append(["TC ID", "Test Step", "Expected Result"])
    second.append(["TC-1", "중복 절차", "중복 결과"])
    workbook.save(path)

    parsed = parse_testcases(path)

    assert len(parsed) == 1
    assert parsed[0].step == "첫 절차"
